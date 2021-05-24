from openpyxl import load_workbook, Workbook
wb = Workbook()
ws = wb.active


wb2 = load_workbook('dane.xlsx')

print(wb2.sheetnames)

c = ws['A1']
print(c)




"""for row in range (1, wd.get_highest_row() + 1):
    a = sheet['A' + str(row)].value"""
